import json
import openpyxl
import datetime

JOB = "0519-LR6"
#JOB = input("What job would you like to input?: ")

wb = openpyxl.load_workbook('QS-100 MASTER SCHEDULE.xlsx', data_only = True)
sheet = wb[JOB]
MAX = 35
#MAX = sheet.max_row

def format_date(d):
    year = str(d.year)
    month = str(d.month).zfill(2)
    day = str(d.day).zfill(2)
    final_date = "{}{}{}".format(month,day,year)
    return(final_date)

#* MO number
#* Drawing number & Description

job_list = []

for row in range(4,MAX+1):
    builder = {}
    for i in range(1,5):
        mo = str(sheet.cell(row=row, column=i).value)
        description_drawing = sheet.cell(row=row, column=i+1).value
        if "None" not in mo:
            break
    builder['Manufacturing Number'] = mo
    if "PKG" in description_drawing:
        description_drawing_split = description_drawing.split(" ")
        builder['Drawing No.'] = description_drawing_split[0]
        builder['Description'] = description_drawing_split[0]
    else:
        description_drawing_split = description_drawing.split(" ")
        builder['Drawing No.'] = description_drawing_split[0]

        try:
            builder['Description'] = description_drawing_split[2]
        except:
            continue

    #* Start date
    start_date = sheet.cell(row=row, column=9).value
    builder['Start Date'] = format_date(start_date)

    #* Finish date
    finish_date = sheet.cell(row=row, column=10).value
    builder['Finish Date'] = format_date(finish_date)

    #* QTY
    qty = str(sheet.cell(row=row, column=11).value)
    builder['Quantity'] = qty
    job_list.append(builder)

#* Save all info into a json file
complete_job = {JOB:job_list}
job_json = json.dumps(complete_job, indent=2)
print(job_json)

with open("manufacturing orders.json", 'w', encoding='utf-8') as f:
    json.dump(complete_job, f, ensure_ascii=False, indent=2)


#TODO - Sort json file in alphabetical order. Or use a range to go in numerical order.
#TODO - Use that info to release MO's
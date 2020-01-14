import json
import openpyxl
import datetime

def format_date(d):
    year = str(d.year)
    month = str(d.month).zfill(2)
    day = str(d.day).zfill(2)
    final_date = "{}{}{}".format(month,day,year)
    return(final_date)

wb = openpyxl.load_workbook('QS-100 MASTER SCHEDULE.xlsx', data_only = True)

jobs = []
job_list = []

for sheet in wb:

    #MAX = 35
    JOB = str(sheet.cell(row=1, column=1).value)
    CUSTOMER = str(sheet.cell(row=3, column=1).value)
    MAX = sheet.max_row

    #* MO number
    #* Drawing number & Description

    
    job_builder = {}

    for row in range(5,MAX+1):
        builder = {}
        builder['Job'] = JOB
        builder['Customer'] = CUSTOMER
        for i in range(1,5):
            mo = str(sheet.cell(row=row, column=i).value)
            description_drawing = sheet.cell(row=row, column=i+1).value
            if "None" not in mo:
                break
        
        builder['Manufacturing Order'] = mo
        #if "PKG" in description_drawing:
            #description_drawing_split = description_drawing.split(" ")
            #builder['Drawing No.'] = description_drawing_split[0]
            #builder['Description'] = description_drawing_split[0]
        #else:
        description_drawing_split = description_drawing.split(" - ")
        builder['Drawing No.'] = description_drawing_split[0]

        try:
            builder['Description'] = description_drawing_split[1]
        except:
            continue

        #* Start date
        start_date = sheet.cell(row=row, column=9).value
        builder['Start Date'] = format_date(start_date)

        #* Finish date
        finish_date = sheet.cell(row=row, column=10).value
        builder['Finish Date'] = format_date(finish_date)

        #* QTY
        qty = sheet.cell(row=row, column=11).value
        builder['Quantity'] = qty
        job_list.append(builder)

    #* Save all info into a json file
    job = {JOB:job_list}
    jobs.append(job)

complete_job = {"MOs": job_list}
    
with open("QS-100 MASTER SCHEDULE.json", 'w', encoding='utf-8') as f:
        json.dump(complete_job, f, ensure_ascii=False, indent=2)

#*Sort json file in alphabetical order. Or use a range to go in numerical order. (I use a range)


job_list = []
with open("QS-100 MASTER SCHEDULE.json") as f:
    data = json.load(f)
    for k in data["MOs"]:
        job = k["Job"]
        if job not in job_list:
            job_list.append(job)

for p in range(0,len(job_list)):
    for i in data["MOs"]:
        job = i["Job"]
        mo = i["Manufacturing Order"]
        if job == job_list[p]:
            print(mo)